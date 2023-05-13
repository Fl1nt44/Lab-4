#!/usr/bin/env python
# coding: utf-8

# # Работа с Excel

# Материалы:
# * Макрушин С.В. Лекция 7: Работа с Excel
# * https://docs.xlwings.org/en/stable/quickstart.html
# * https://nbviewer.jupyter.org/github/pybokeh/jupyter_notebooks/blob/master/xlwings/Excel_Formatting.ipynb#search_text
# 

# ## Задачи для совместного разбора

# 1. На листе "Рецептура" файла `себестоимостьА_в1.xlsx` для области "Пшеничный хлеб" рассчитать себестоимость всех видов продукции.

# 2. Результаты расчетов 1.1 сохранить в отдельном столбце области "Пшеничный хлеб"

# 3. Приблизить форматирование столбца, добавленного в задаче 2 к оформлению всей области.

# 4. Выполнить 3 с помощью "протягиваемых" формул.

# ## Лабораторная работа 7.1

# 1. Загрузите данные из файлов `reviews_sample.csv` (__ЛР2__) и `recipes_sample.csv` (__ЛР5__) в виде `pd.DataFrame`. Обратите внимание на корректное считывание столбца(ов) с индексами. Оставьте в таблице с рецептами следующие столбцы: `id`, `name`, `minutes`, `submitted`, `description`, `n_ingredients`

# In[4]:


import pandas as pd

# Загрузка данных из файлов
reviews = pd.read_csv('data/reviews_sample.csv', index_col=0)
recipes = pd.read_csv('data/recipes_sample.csv', index_col='id')

# Оставляем только нужные столбцы в таблице с рецептами
recipes = recipes[['name', 'minutes', 'submitted', 'description', 'n_ingredients']]


# 2. Случайным образом выберите 5% строк из каждой таблицы и сохраните две таблицы на разные листы в один файл `recipes.xlsx`. Дайте листам названия "Рецепты" и "Отзывы", соответствующие содержанию таблиц. 

# In[5]:


# Задание 2
recipes_sample = recipes.sample(frac=0.05, random_state=42)
reviews_sample = reviews.sample(frac=0.05, random_state=42)

with pd.ExcelWriter('recipes.xlsx') as writer:
    recipes_sample.to_excel(writer, sheet_name='Рецепты', index=True)
    reviews_sample.to_excel(writer, sheet_name='Отзывы', index=False)


# 3. Используя `xlwings`, добавьте на лист `Рецепты` столбец `seconds_assign`, показывающий время выполнения рецепта в секундах. Выполните задание при помощи присваивания массива значений диапазону ячеек.

# In[6]:


import pandas as pd
import xlwings as xw

# Загрузка данных из файла
recipes = pd.read_excel('recipes.xlsx', sheet_name='Рецепты', index_col=0)

# Вычисление времени выполнения рецепта в секундах
recipes['seconds_assign'] = recipes['minutes'] * 60

# Открытие файла с помощью xlwings
wb = xw.Book('recipes.xlsx')
sheet = wb.sheets['Рецепты']

# Запись данных в столбец seconds_assign
sheet.range('F1').value = 'seconds_assign'
sheet.range('F2').options(transpose=True).value = recipes['seconds_assign']

# Сохранение изменений и закрытие файла
wb.save()
wb.close()


# 4. Используя `xlwings`, добавьте на лист `Рецепты` столбец `seconds_formula`, показывающий время выполнения рецепта в секундах. Выполните задание при помощи формул Excel.

# In[9]:


# Открываем файл Excel
wb = xw.Book('recipes.xlsx')

# Получаем лист Рецепты
sheet = wb.sheets['Рецепты']

# Получаем данные из листа Рецепты в виде pandas DataFrame
df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

# Вычисляем время выполнения рецепта в секундах
df['seconds_formula'] = df['minutes'] * 60

# Записываем данные обратно на лист Рецепты
sheet.clear_contents()
sheet.range('A1').value = df

# Сохраняем файл Excel
wb.save()
wb.close()


# 5. Сделайте названия всех добавленных столбцов полужирными и выровняйте по центру ячейки.

# In[18]:


import xlwings as xw
import xlwings as xw

# Открываем файл Excel
wb = xw.Book('recipes.xlsx')

# Получаем лист Рецепты
sheet = wb.sheets['Рецепты']

# Делаем названия всех добавленных столбцов полужирными и выравниваем их по центру ячеек
for col in sheet.range('A1').expand('right').columns:
    col.api.Font.Bold = True
    col.api.HorizontalAlignment = xw.constants.HAlign.center

# Сохраняем файл Excel
wb.save()
wb.close()


# 6. Раскрасьте ячейки столбца `minutes` в соответствии со следующим правилом: если рецепт выполняется быстрее 5 минут, то цвет - зеленый; от 5 до 10 минут - жёлтый; и больше 10 - красный.

# In[19]:


# Открываем файл Excel
wb = xw.Book('recipes.xlsx')

# Получаем доступ к листу "Рецепты"
sheet = wb.sheets['Рецепты']

# Загружаем данные из листа в pandas DataFrame
df = sheet.used_range.options(pd.DataFrame, header=1, index=False).value

# Функция для определения цвета ячейки в зависимости от значения
def get_color(val):
    if val < 5:
        return 'green'
    elif val >= 5 and val < 10:
        return 'yellow'
    else:
        return 'red'

# Применяем функцию к столбцу "minutes" и сохраняем результат в новый столбец "color"
df['color'] = df['minutes'].apply(get_color)

# Получаем доступ к столбцу "color" и устанавливаем цвет ячеек в соответствии с его значениями
color_range = sheet.range((2, df.columns.get_loc('color') + 1), (df.shape[0] + 1, df.columns.get_loc('color') + 1))
color_range.color = [get_color(val) for val in df['minutes']]

# Сохраняем изменения в файл Excel
wb.save()


# 7. Добавьте на лист `Рецепты`  столбец `n_reviews`, содержащий кол-во отзывов для этого рецепта. Выполните задание при помощи формул Excel.

# In[20]:


import pandas as pd
import xlwings as xw

# Открываем файл Excel
wb = xw.Book('recipes.xlsx')

# Получаем доступ к листу "Рецепты"
sheet = wb.sheets['Рецепты']

# Получаем диапазон ячеек со значениями столбца "reviews"
reviews_range = sheet.range((2, sheet.range('A1').current_region.last_cell.column), (sheet.range('A1').current_region.last_cell.row, sheet.range('A1').current_region.last_cell.column))

# Добавляем новый столбец "n_reviews" и заполняем его формулой COUNTIF
n_reviews_col = sheet.range((1, sheet.range('A1').current_region.last_cell.column + 1), (sheet.range('A1').current_region.last_cell.row, sheet.range('A1').current_region.last_cell.column + 1))
n_reviews_col.value = [['n_reviews']]
n_reviews_range = sheet.range((2, sheet.range('A1').current_region.last_cell.column + 1), (sheet.range('A1').current_region.last_cell.row, sheet.range('A1').current_region.last_cell.column + 1))
n_reviews_range.formula = f'=COUNTIF({reviews_range.address}, A2)'

# Сохраняем изменения в файл Excel
wb.save()


# ## Лабораторная работа 7.2

# 8. Напишите функцию `validate()`, которая проверяет соответствие всех строк из листа `Отзывы` следующим правилам:
#     * Рейтинг - это число от 0 до 5 включительно
#     * Соответствующий рецепт имеется на листе `Рецепты`
#     
# В случае несоответствия этим правилам, выделите строку красным цветом

# In[13]:


import pandas as pd
from openpyxl.styles import PatternFill

# Загрузка данных из файла Excel
df_reviews = pd.read_excel('recipes.xlsx', sheet_name='Отзывы')
df_recipes = pd.read_excel('recipes.xlsx', sheet_name='Рецепты')

# Настройка стиля для выделения красным цветом
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# Определение функции для проверки соответствия строк правилам
def validate():
    df_reviews = df_recipes.astype(str)
    for index, row in df_reviews.iterrows():
        # Проверка соответствия рейтинга правилам
        if not (0.0 <= row['rating'] <= 5.0):
            # Выделение ячейки с рейтингом красным цветом, если рейтинг неверен
            df_reviews.loc[index, 'rating'] = df_reviews.loc[index, 'rating'].apply(lambda x: f'background-color: {red_fill.fill_type}; color: {red_fill.start_color}' if not (0 <= x <= 5) else '')
        # Проверка наличия рецепта на листе Рецепты
        if row['recipe_id'] not in df_recipes['id'].values:
            # Выделение ячейки с ID рецепта красным цветом, если рецепта нет на листе Рецепты
            df_reviews.loc[index, 'id'] = df_reviews.loc[index, 'id'].apply(lambda x: f'background-color: {red_fill.fill_type}; color: {red_fill.start_color}' if x not in df_recipes['id'].values else '')

# Вызов функции для проверки соответствия строк правилам
validate()

# Сохранение изменений в файл Excel
with pd.ExcelWriter('recipes.xlsx', engine='openpyxl') as writer:
    writer.book = load_workbook('recipes.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    df_reviews.to_excel(writer, sheet_name='Отзывы', index=False)


# 9. В файле `recipes_model.csv` находится модель данных предметной области "рецепты". При помощи пакета `csv` считайте эти данные. При помощи пакета `xlwings` запишите данные на лист `Модель` книги `recipes_model.xlsx`, начиная с ячейки `A2`, не используя циклы. Сделайте скриншот текущего состояния листа и прикрепите в ячейку ноутбука. 

# 10. При помощи пакета `xlwings` добавьте в столбец J формулу для описания столбца на языке SQL. Формула должна реализовывать следующую логику:
# 
#     1\. в начале строки идут значения из столбцов В и C (значение столбца С приведено к верхнему регистру), разделенные пробелом
#     
#     2\. далее идут слова на основе столбца "Ключ"
#         2.1 если в столбце "Ключ" указано значение "PK", то дальше через пробел идет ключевое слово "PRIMARY KEY"
#         2.2 если в столбце "Ключ" указано значение "FK", то дальше через пробел идет ключевое слово "REFERENCES", затем значения столбцов H и I в формате "название_таблицы(название_столбца)"
#         
#     3\. если в столбце "Обязательно к заполнению" указано значение "Y" и в столбце "Ключ" указано не "PK", то дальше через пробел идет ключевое слово "NOT NULL".
# 
# Заполните этой формулой необходимое количество строк, используя "протягивание". Количество строк для протягивания определите на основе данных.
# 
# Сделайте скриншот текущего состояния листа и прикрепите в ячейку ноутбука.

# 11. При помощи пакета `xlwings` измените стилизацию листа `Модель`.
# * для заголовков добавьте заливку цвета `00ccff`
# * примените автоподбор ширины столбца;
# * сделайте шрифт заголовков полужирным;
# * добавьте таблице автофильтр.
# 
# Сделайте скриншот текущего состояния листа и прикрепите в ячейку ноутбука.

# 12. Посчитайте количество атрибутов для каждой из сущностей. Создайте лист `Статистика` и запишите в него результат группировки, начиная с ячейки "А1". Визуализируйте полученный результат при помощи столбчатой диаграммы. Сохраните полученную визуализацию на лист `Статистика`, начиная с ячейки "E2".  Сделайте скриншот листа `Статистика` и прикрепите в ячейку ноутбука.
# 
# * Вы можете воспользоваться методами для визуализации, которые поставляются вместе с объектами `pandas` (см. https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.plot) 
